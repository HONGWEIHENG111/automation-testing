import os
import time
import queue
import re
from datetime import datetime
import threading
import openpyxl
import keyboard
from webdriver_manager.chrome import ChromeDriverManager
from utils.logger import setup_global_logger
from utils.config import USERNAME, PASSWORD, HOME_URL, INPUT_EXCEL_PATH
from utils.common import (
    show_authorship, close_popups, get_test_data_from_excel, init_browser, perform_login
)
from tools.summary_generator import generate_summary_csv

# ================= 引入 Core 模块 =================
from utils.signals import STOP_EVENT
from utils.models import TaskInput
from utils.pipeline import process_single_task


def listen_for_hotkey():
    try:
        keyboard.add_hotkey('ctrl+alt+h', show_authorship)
        keyboard.wait('ctrl+q')
        STOP_EVENT.set()
        print("\n\n🛑 [紧急刹车] 侦测到 Ctrl+Q 键按下！脚本将在当前步骤跳出...\n")
    except ImportError:
        print("⚠️ 缺少 keyboard 模块，无法使用快捷键终止功能。")
    except Exception as e:
        print(f"⚠️ 热键监听启动失败 (可能缺少管理员权限)，将禁用 Ctrl+Q 功能: {e}")

def excel_writer_worker(excel_path, q, log_func=print):
    """专职 Excel 写入后台线程（带非法字符过滤）"""
    import openpyxl
    try:
        wb = openpyxl.load_workbook(excel_path)
        ws = wb.active
    except Exception as e:
        log_func(f"❌ [写线程致命错误] 无法加载 Excel: {e}")
        return

    while True:
        item = q.get()
        if item == "STOP":
            try: wb.save(excel_path)
            except: pass
            break

        task, result = item
        try:
            row_data = [
                task.index + 1, task.question_text, task.filename if task.filename else "",
                result.crash_reason, result.tester_expectation, task.target_language if task.target_language else "N/A",
                result.input_language, result.output_language, result.language_status,
                result.answer_text, result.shared_link, result.evaluation_text,
                result.actual_agent_used, result.reference_link, result.document_contain,
                result.prep_time, result.comp_time, result.timeout_status
            ]
            target_row = task.index + 2
            for col_index, value in enumerate(row_data, start=1):
                # 🛡️ 核心防御：过滤 Excel 不支持的非法控制字符
                if isinstance(value, str):
                    value = re.sub(r'[\x00-\x08\x0b\x0c\x0e-\x1f]', '', value)
                ws.cell(row=target_row, column=col_index, value=value)

            save_success = False
            for retry in range(3):
                try:
                    wb.save(excel_path)
                    save_success = True
                    break
                except PermissionError:
                    time.sleep(3)
            if not save_success:
                log_func(f"❌ [致命错误] 多次尝试保存失败，第 {task.index + 1} 行结果丢失！")
        except Exception as e:
            log_func(f"❌ [写线程异常] 写入行 {task.index + 1} 失败: {e}")
        finally:
            q.task_done()

def run_automation():
    setup_global_logger()
    print("📥 正在初始化並緩存瀏覽器驅動...")
    cached_driver_path = ChromeDriverManager().install()
    threading.Thread(target=listen_for_hotkey, daemon=True).start()

    project_dir = os.path.dirname(os.path.abspath(__file__))
    test_dir = os.path.join(project_dir, "test")

    if not os.path.exists(test_dir):
        print(f"⚠️ [警告] 找不到测试文件夹: {test_dir}，已自动创建。若本次测试包含文件上传，请确保文件放入该目录。")
        os.makedirs(test_dir, exist_ok=True)

    questions, selected_agents, target_filenames, selected_languages = get_test_data_from_excel(INPUT_EXCEL_PATH)
    if not questions:
        print("未提取到任何问题，程序终止。")
        return
    # ================= 文件数量校验 =================
    valid_extensions = ('.pdf', '.docx', '.csv', '.txt')
    test_files = [f for f in os.listdir(test_dir) if f.lower().endswith(valid_extensions)]
    expected_file_count = len([f for f in target_filenames if str(f).strip()])
    if len(test_files) != expected_file_count:
        print(f"⚠️ 文件夹中的有效文件数量 ({len(test_files)}) 与 Excel 中的问题数量 ({len(questions)}) 不匹配！")
    # ================= 核心修改：打包 TaskInput =================
    task_inputs = []
    for i in range(len(questions)):
        task_inputs.append(TaskInput(
            index=i,
            question_text=questions[i],
            target_agent=selected_agents[i],
            filename=target_filenames[i],
            target_language=selected_languages[i]
        ))

    # ================= 初始化 Excel 文件 =================
    base_testcase_name = os.path.splitext(os.path.basename(INPUT_EXCEL_PATH))[0]
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    dynamic_filename = f"evaluation_results_{base_testcase_name}_{timestamp}.xlsx"
    excel_dir = os.path.join(project_dir, "Evaluation_Results")
    os.makedirs(excel_dir, exist_ok=True)
    excel_path = os.path.join(excel_dir, dynamic_filename)

    if not os.path.exists(excel_path):
        wb = openpyxl.Workbook()
        wb.properties.creator = "Henry HONG "
        wb.properties.description = "Authored by Henry HONG. "
        ws = wb.active
        ws.title = "Evaluation Results"
        ws.append(["label", "Request", "filename", "Crash", "Tester Expectation", "Selected Language", "Input Language",
                   "Output Language", "Language Overall Status", "answer", "shared link", "DeepSeek评价内容",
                   "Selected agent", "Reference Link", "Document Contain[1][2][3]", "Preparation Time",
                   "Completion Time", "Timeout_States"])
        wb.save(excel_path)
        print(f"📊 已创建评价结果 Excel 文件: {excel_path}")

    result_queue = queue.Queue()
    writer_thread = threading.Thread(target=excel_writer_worker, args=(excel_path, result_queue, print), daemon=True)
    writer_thread.start()

    # ================= 启动专属浏览器 =================
    try:
        driver = init_browser(chrome_driver_path=cached_driver_path)
        driver.get(HOME_URL)
        time.sleep(1)
        perform_login(driver, USERNAME, PASSWORD)
    except Exception as e:
        print(f"❌ 浏览器启动或登录失败，程序终止: {e}")
        try:
            driver.quit()  # 如果账号密码错了进不去，自动把浏览器关掉
        except:
            pass
        return

    # ================= 顺序循环处理任务 =================
    try:
        for task in task_inputs:
            if STOP_EVENT.is_set():
                print("🛑 收到中止指令，停止处理后续文件。")
                break

            # 1. 暴力兜底：清理遗留弹窗
            try:
                from selenium.webdriver.support.ui import WebDriverWait
                from selenium.webdriver.support import expected_conditions as EC
                WebDriverWait(driver, 0.5).until(EC.alert_is_present())
                driver.switch_to.alert.accept()
            except Exception:
                pass

            # 2. 返回主页
            try:
                driver.get(HOME_URL)
                time.sleep(1.5)
            except Exception:
                print("⚠️ 无法返回主页，尝试继续执行...")

            # 3. 核心流水线调用
            print(f"\n--- 进度: {task.index + 1}/{len(task_inputs)} ---")
            status = process_single_task(
                task=task,
                driver=driver,
                test_dir=test_dir,
                result_queue=result_queue,
                stop_event=STOP_EVENT,
                log_func=print
            )

            if status == "BROWSER_CLOSED":
                break

            # 4. 内存保护机制：每 50 次强行释放并重启 (100% 对齐原逻辑)
            if (task.index + 1) % 50 == 0 and not STOP_EVENT.is_set():
                print(f"\n[{task.index + 1}] 🔄 触发内存保护：已连续运行 50 次，正在重启浏览器释放内存...")
                try:
                    driver.quit()
                except:
                    pass

                try:
                    driver = init_browser(chrome_driver_path=cached_driver_path)
                    driver.get(HOME_URL)
                    time.sleep(2)
                    perform_login(driver, USERNAME, PASSWORD)
                    print(f"[{task.index + 1}] ✅ 内存释放完毕，新浏览器登录成功，继续后续任务...")
                except Exception as e:
                    print(f"[{task.index + 1}] ❌ 重启浏览器失败: {e}，终止任务。")
                    break
    except KeyboardInterrupt:
        print("\n🛑 侦测到系统强行中断信号 (Ctrl+C)！脚本被迫停止！")
        STOP_EVENT.set()

    except Exception as e:
        print(f"\n❌ 主脚本发生未预期的致命错误: {e}")
    finally:
    # ================= 生成报告与清理 =================
        print("💾 正在等待残余数据写入 Excel...")
        if 'result_queue' in locals():
            result_queue.put("STOP")
            writer_thread.join(timeout=10)
        print("📊 正在生成最终的 Summary 报告...")
        output_csv = os.path.join(project_dir, "Summaries", f"Summary_{base_testcase_name}_{timestamp}.csv")
        os.makedirs(os.path.dirname(output_csv), exist_ok=True)
        try:
            generate_summary_csv(excel_path, output_csv)
            print(f"✅ 汇总报告已生成: {output_csv}")
        except Exception as e:
            print(f"⚠️ 生成汇总报告时发生异常: {e}")

        print("👉 自动化控制权已释放，浏览器将保持开启状态。")
        while True:
            if STOP_EVENT.is_set(): break
            try:
                _ = driver.window_handles
                time.sleep(1)
            except Exception:
                print("\n👋 侦测到浏览器已被手动关闭，控制台自动退出！")
                break


if __name__ == "__main__":
    run_automation()
