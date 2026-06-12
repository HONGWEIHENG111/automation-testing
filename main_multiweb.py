import os
import time
from datetime import datetime
from docx import Document
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.edge.service import Service as EdgeService
from selenium.webdriver.edge.options import Options as EdgeOptions
from webdriver_manager.microsoft import EdgeChromiumDriverManager
import PyPDF2
import base64
import openpyxl
import keyboard
import threading
import re
from selenium.webdriver.chrome.options import Options
from tools.state_manager import execute_state, handle_post_send
from tools.llm_evaluator import evaluate_with_deepseek
from tools.summary_generator import generate_summary_csv
from tools.agent_mapping import AGENT_NAME_MAPPING_TC, TC_TO_ENG_MAPPING
import difflib

from config import (
    INPUT_EXCEL_PATH,
    URL_1, ENV_NAME_1, USERNAME_1, PASSWORD_1,
    URL_2, ENV_NAME_2, USERNAME_2, PASSWORD_2
)
from threading import Lock
PRINT_LOCK = Lock()

def safe_print(*args, **kwargs):
    """线程安全的打印函数，防止双网页并发时控制台字符错乱"""
    with PRINT_LOCK:
        print(*args, **kwargs)
# 动态生成所有合法的 Agent 名称全局列表（包含所有中英文标准名称）
ALL_VALID_AGENTS = list(AGENT_NAME_MAPPING_TC.keys())
for tc_list in AGENT_NAME_MAPPING_TC.values():
    ALL_VALID_AGENTS.extend(tc_list)
ALL_VALID_AGENTS = list(set(ALL_VALID_AGENTS))  # 去重备用

STOP_SCRIPT = False


def listen_for_hotkey():
    """在后台运行的线程，专门监听终止键"""
    global STOP_SCRIPT
    keyboard.add_hotkey('ctrl+alt+h', show_authorship)
    # 阻塞等待按下 Ctrl+C
    keyboard.wait('ctrl+c')
    STOP_SCRIPT = True
    print("\n\n🛑 [紧急刹车] 侦测到 Ctrl+C 键按下！脚本将在当前步骤跳出，并释放浏览器控制权...\n")


def smart_sleep(seconds):
    """可随时被紧急刹车打断的睡眠函数"""
    for _ in range(int(seconds * 10)):
        if STOP_SCRIPT:
            break
        time.sleep(0.1)


def close_popups(driver, env_name):
    """专门用于检测并关闭页面上的干扰弹窗"""
    safe_print(f"[{env_name}] 🔍 正在执行弹窗清理扫雷...")
    try:
        skip_btn = WebDriverWait(driver, 1).until(
            EC.presence_of_element_located((By.ID, "gen-tour-welcome-skip"))
        )
        driver.execute_script("arguments[0].click();", skip_btn)
        safe_print(f"[{env_name}] ✅ 已成功跳过 Welcome 导览弹窗")
        time.sleep(0.5)
    except Exception:
        pass

    try:
        dont_show_label = WebDriverWait(driver, 1).until(
            EC.presence_of_element_located((By.XPATH, "//label[@for='dontShowAgain']"))
        )
        driver.execute_script("arguments[0].click();", dont_show_label)
        safe_print(f"[{env_name}] ✅ 已勾选 'Don't show this again'")
        time.sleep(0.5)

        close_btn = driver.find_element(By.CLASS_NAME, "credit-popup-close")
        driver.execute_script("arguments[0].click();", close_btn)
        safe_print(f"[{env_name}] ✅ 已成功关闭 News 弹窗")
    except Exception:
        pass

def get_test_data_from_excel(excel_path):
    """从 Excel 中智能寻找 'Request' 和 'Selected Agent'，并提取其正下方的所有内容"""
    questions = []
    selected_agents = []
    filenames = []
    selected_languages = []

    if not os.path.exists(excel_path):
        print(f"[致命错误] 找不到 Excel 文件: {excel_path}")
        return questions, selected_agents, filenames, selected_languages

    try:
        wb = openpyxl.load_workbook(excel_path)
        ws = wb.active

        request_col = None
        agent_col = None
        filename_col = None
        language_col = None
        header_row = None

        # 1. 遍历表格，寻找表头
        for row in ws.iter_rows():
            for cell in row:
                if cell.value:
                    val = str(cell.value).strip().lower()
                    if val == "request":
                        header_row = cell.row
                        request_col = cell.column
                    elif val == "selected agent":
                        header_row = cell.row  # 假设它们在同一行
                        agent_col = cell.column
                    elif val in ["filename", "file name", "file_name"]:  # 兼容不同写法
                        header_row = cell.row
                        filename_col = cell.column
                    else:
                        # 标准和常见变体词库
                        lang_keywords = ["selected language", "selected_language", "language", "lang", "目标语言",
                                         "测试语言", "ui语言", "ui language"]
                        is_language_column = False

                        # 1. 精确词库匹配
                        if val in lang_keywords:
                            is_language_column = True

                        # 2. 子串包含匹配（只要包含核心词根就命中）
                        elif "language" in val or "lang" in val or "语言" in val:
                            is_language_column = True

                        # 3. 模糊相似度匹配（容忍测试人员手误拼错，如 "langauge"）
                        else:
                            matches = difflib.get_close_matches(val, lang_keywords, n=1, cutoff=0.6)
                            if matches:
                                is_language_column = True

                        # 如果通过以上任一防线确认是语言列，则记录列号
                        if is_language_column:
                            header_row = cell.row
                            language_col = cell.column
            # 只要发现了 Request 或 Filename 任意一个核心锚点，就确认是表头并跳出扫描。
            # 这样既不会被可能在第一行出现的随机备注（如 "默认Language"）骗到，也能完整扫描完这一行的所有列。
            if request_col is not None or filename_col is not None:
                break

        if not request_col and not filename_col:
            print(f"[致命错误] 在 Excel({excel_path}) 中找不到 'Request' 也找不到 'Filename' 表头！无法提供任何测试输入。")
            return questions, selected_agents, filenames, selected_languages

        if not filename_col:
            print(f"[降级提示] 找不到 'Filename' 表头，将自动降级为【纯文本模式】运行。")

        if not agent_col:
            print(f"[降级提示] 找不到 'Selected Agent' 表头，将自动降级为【状态 2 (通用自动模式)】。")

        if not language_col:
            print(f"[降级提示] 找不到 'Language' 表头，将自动使用【英文 (默认)】。")

        for row in range(header_row + 1, ws.max_row + 1):
            req_val = ws.cell(row=row, column=request_col).value if request_col else ""
            file_val = ws.cell(row=row, column=filename_col).value if filename_col else ""

            req_str = str(req_val).strip() if req_val else ""
            file_str = str(file_val).strip() if file_val else ""

            # 只要 Request 不为空 或者 Filename 不为空，就认为这是一条有效数据
            if req_str or file_str:
                questions.append(req_str)
                filenames.append(file_str)
                # 读取对应的 Agent
                if agent_col:
                    agent_val = ws.cell(row=row, column=agent_col).value
                    selected_agents.append(str(agent_val).strip() if agent_val else "")
                else:
                    selected_agents.append("")

                # 读取对应的 Language
                if language_col:
                    lang_val = ws.cell(row=row, column=language_col).value
                    selected_languages.append(str(lang_val).strip() if lang_val else "")
                else:
                    selected_languages.append("")
    except Exception as e:
        print(f"[格式错误] 读取 Excel 文件失败: {e}")

    return questions, selected_agents, filenames, selected_languages

def show_authorship():
    secret_msg = base64.b64decode(b'Q29kZSBBdXRob3I6IEhlbnJ5IEhPTkc=').decode('utf-8')
    print(f"\n\n====================================")
    print(f"🌟 {secret_msg} 🌟")
    print(f"====================================\n\n")


def read_file_content(file_path, env_name=""):
    ext = os.path.splitext(file_path)[1].lower()
    content = ""
    prefix = f"[{env_name}] " if env_name else ""
    try:
        if ext in ['.txt', '.csv']:
            try:
                with open(file_path, 'r', encoding='utf-8') as f:
                    content = f.read()
            except UnicodeDecodeError:
                with open(file_path, 'r', encoding='gbk') as f:
                    content = f.read()
        elif ext == '.docx':
            doc = Document(file_path)
            content = "\n".join([para.text for para in doc.paragraphs if para.text.strip()])
        elif ext == '.pdf':
            with open(file_path, 'rb') as f:
                reader = PyPDF2.PdfReader(f)
                content = "\n".join([page.extract_text() for page in reader.pages if page.extract_text()])
    except Exception as e:
        safe_print(f"{prefix}❌ 读取文件 {os.path.basename(file_path)} 内容时失败: {e}")
    return content


# 注入用户名密码作为函数参数
def run_automation(home_url, env_name, username, password, test_data, chrome_driver_path):
    global STOP_SCRIPT

    CURRENT_STATE = "4"
    project_dir = os.path.dirname(os.path.abspath(__file__))
    test_dir = os.path.join(project_dir, "test")

    if not os.path.exists(test_dir):
        safe_print(f"[{env_name}] [致命错误] 找不到测试文件夹: {test_dir}")
        return

    questions, selected_agents, target_filenames, selected_languages = test_data
    if not questions:
        safe_print(f"[{env_name}] 未提取到任何问题，程序终止。")
        return
    if len(questions) != len(selected_agents):
        safe_print(f"[{env_name}] [致命错误] 提取到的问题数量与 Agent 数量不一致！请检查 Excel 格式。")
        return

    valid_extensions = ('.pdf', '.docx', '.csv', '.txt')
    test_files = [f for f in os.listdir(test_dir) if f.lower().endswith(valid_extensions)]
    test_files.sort()

    if len(test_files) != len(questions):
        safe_print(f"[{env_name}] [警告] 文件夹中的文件数量 ({len(test_files)}) 与 Excel 中的问题数量不匹配！")

    try:
        safe_print(f"[{env_name}] 🔍 尝试启动首选浏览器：Google Chrome...")
        chrome_options = Options()
        chrome_options.add_experimental_option("detach", True)
        chrome_options.add_argument('--ignore-certificate-errors')
        # 禁止浏览器后台休眠和节流
        chrome_options.add_argument('--disable-background-timer-throttling')
        chrome_options.add_argument('--disable-backgrounding-occluded-windows')
        chrome_options.add_argument('--disable-renderer-backgrounding')
        # chrome_options.add_argument('--headless=new') #当完全没有问题之后可以使用这个无头模式，不用盯着界面看，节省70%内存，可以调高worknum
        # chrome_options.add_argument('--disable-gpu')
        driver = webdriver.Chrome(service=Service(chrome_driver_path), options=chrome_options)
        safe_print(f"[{env_name}] ✅ 成功启动 Google Chrome！")
    except Exception as e_chrome:
        safe_print(f"[{env_name}] ⚠️ Chrome 启动失败: {e_chrome}")
        safe_print(f"[{env_name}] 🔄 正在尝试启动备用浏览器：Microsoft Edge...")
        try:
            edge_options = EdgeOptions()
            edge_options.add_experimental_option("detach", True)
            edge_options.add_argument('--ignore-certificate-errors')
            driver = webdriver.Edge(service=EdgeService(EdgeChromiumDriverManager().install()), options=edge_options)
            safe_print(f"[{env_name}] ✅ 成功启动 Microsoft Edge！")
        except Exception as e_edge:
            safe_print(f"[{env_name}] ❌ [致命错误] Chrome 和 Edge 均启动失败，请检查浏览器安装环境！\n错误信息: {e_edge}")
            return

    driver.get(home_url)
    time.sleep(1)

    # ================= 新自动登录逻辑 (JS 强行注入版本对齐) =================
    safe_print(f"[{env_name}] 🔑 正在执行自动登录...")
    try:
        login_nav = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.ID, "nav-login"))
        )
        driver.execute_script("arguments[0].click();", login_nav)

        username_input = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.ID, "sso-username"))
        )
        password_input = driver.find_element(By.ID, "sso-password")

        # 使用传入的对应环境账号密码
        driver.execute_script("arguments[0].value = arguments[1];", username_input, username)
        driver.execute_script("arguments[0].dispatchEvent(new Event('input', { bubbles: true }));", username_input)

        driver.execute_script("arguments[0].value = arguments[1];", password_input, password)
        driver.execute_script("arguments[0].dispatchEvent(new Event('input', { bubbles: true }));", password_input)

        submit_btn = driver.find_element(By.CLASS_NAME, "sso-submit-btn")
        driver.execute_script("arguments[0].click();", submit_btn)

        close_popups(driver, env_name)
        safe_print(f"[{env_name}] ✅ 登录动作已提交，等待页面刷新...")
        time.sleep(1)
        close_popups(driver, env_name)
    except Exception as e:
        safe_print(f"[{env_name}] ❌ [登录失败] 请检查页面是否卡顿或元素发生变化: {e}")
        driver.quit()
        return

    # ================= 初始化 Excel 结果文件 =================
    base_testcase_name = os.path.splitext(os.path.basename(INPUT_EXCEL_PATH))[0]
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    dynamic_filename = f"evaluation_results_{env_name}_{base_testcase_name}_{timestamp}.xlsx"
    excel_dir = os.path.join(project_dir, f"Evaluation_Results_{env_name}")
    os.makedirs(excel_dir, exist_ok=True)
    excel_path = os.path.join(excel_dir, dynamic_filename)

    if not os.path.exists(excel_path):
        wb = openpyxl.Workbook()
        wb.properties.creator = "Henry HONG "
        wb.properties.description = "Authored by Henry HONG. "
        ws = wb.active
        ws.title = "Evaluation Results"
        ws.append(["label", "Request", "filename", "Crash", "Tester Expectation", "Selected Language", "Input Language",
                   "Output Language", "Language Overall Status", "answer", "shared link", "DeepSeek评价内容",
                   "Selected agent", "Reference Link", "Document Contain[1][2][3]", "Preparation Time",
                   "Completion Time", "Timeout_States"])
        wb.save(excel_path)
        safe_print(f"[{env_name}] 📊 已创建评价结果 Excel 文件: {excel_path}")

    # ================= 遍历执行问题 =================
    for i in range(len(questions)):
        if STOP_SCRIPT:
            safe_print(f"[{env_name}] 🛑 收到中止指令，停止处理后续文件。")
            break

        timeout_status = "No"
        try:
            _ = driver.window_handles
        except Exception:
            safe_print(f"\n[{env_name}] 🚨 侦测到浏览器窗口已被手动关闭，脚本终止运行！")
            break

        question_text = questions[i]
        target_agent = selected_agents[i]
        filename = target_filenames[i]
        target_language = selected_languages[i]

        actual_file_name = None
        file_path = None
        has_file_input = False

        if filename and str(filename).strip():
            target_base_name = str(filename).strip().lower()
            for f in os.listdir(test_dir):
                base_name_in_folder, ext = os.path.splitext(f)
                if base_name_in_folder.strip().lower() == target_base_name:
                    actual_file_name = f
                    file_path = os.path.join(test_dir, actual_file_name)
                    has_file_input = True
                    break

            if not has_file_input:
                if question_text and str(question_text).strip():
                    safe_print(f"\n[{env_name}] ⚠️ [降级运行] 找不到文件 '{filename}'，降级为仅发送问题的纯文本模式！")
                else:
                    safe_print(f"\n[{env_name}] ❌ [跳过任务] 找不到文件 '{filename}' 且没有 Request，跳过此条任务！")
                    continue
        else:
            if not question_text or not str(question_text).strip():
                safe_print(f"\n[{env_name}] ❌ [跳过任务] Request 和 Filename 同时为空，跳过此条任务！")
                continue

        is_text_only = not has_file_input

        if target_agent and str(target_agent).strip():
            CURRENT_STATE = "4"
        else:
            CURRENT_STATE = "2"

        target_agent_raw = str(target_agent).strip() if target_agent else ""

        if target_agent_raw and target_agent_raw not in ALL_VALID_AGENTS:
            lower_agents_mapping = {agent.lower(): agent for agent in ALL_VALID_AGENTS}
            target_agent_lower = target_agent_raw.lower()

            # 第一步防御：无视大小写的精确匹配
            if target_agent_lower in lower_agents_mapping:
                correct_name = lower_agents_mapping[target_agent_lower]
                safe_print(
                    f"[{env_name}] [{i + 1}] 🔧 [大小写修正]: 侦测到大小写不标准，纠正 '{target_agent_raw}' 为 '{correct_name}'")
                target_agent_raw = correct_name
            else:
                # 第二步防御：模糊拼写纠错
                matches = difflib.get_close_matches(target_agent_lower, list(lower_agents_mapping.keys()), n=1,
                                                    cutoff=0.6)
                if matches:
                    correct_name = lower_agents_mapping[matches[0]]
                    safe_print(
                        f"[{env_name}] [{i + 1}] 🔧 [智能纠错]: 侦测到拼写错误，纠正 '{target_agent_raw}' 为 '{correct_name}'")
                    target_agent_raw = correct_name
                else:
                    safe_print(
                        f"[{env_name}] [{i + 1}] ⚠️ [降级警告]: 无法识别 Agent '{target_agent_raw}'，自动降级为 State 2！")
                    target_agent_raw = ""
                    CURRENT_STATE = "2"

        tc_keywords = ["繁中", "繁体", "繁體", "traditional chinese", "tc", "zh-tw"]
        is_tc_ui = target_language and any(keyword in str(target_language).lower() for keyword in tc_keywords)
        search_candidates = []
        if target_agent_raw:
            if is_tc_ui:
                if target_agent_raw in AGENT_NAME_MAPPING_TC:
                    search_candidates = AGENT_NAME_MAPPING_TC[target_agent_raw]
                    safe_print(
                        f"[{env_name}] 🔄 [Agent 转换]: 繁中 UI 匹配到英文输入，载入候选列表 -> {search_candidates}")
                else:
                    search_candidates = [target_agent_raw]
                    safe_print(
                        f"[{env_name}] 🎯 [Agent 保持]: 繁中 UI 匹配到繁中输入，直接搜索 -> '{target_agent_raw}'")
            else:
                if target_agent_raw in TC_TO_ENG_MAPPING:
                    search_candidates = [TC_TO_ENG_MAPPING[target_agent_raw]]
                    safe_print(
                        f"[{env_name}] 🔄 [Agent 转换]: 英文 UI 匹配到繁中输入，已自动转为英文 -> '{search_candidates[0]}'")
                else:
                    search_candidates = [target_agent_raw]
                    safe_print(f"[{env_name}] 🎯 [Agent 保持]: 英文 UI 匹配到英文输入，直接搜索 -> '{target_agent_raw}'")
        safe_print(f"\n--- [{env_name}] 进度: {i + 1}/{len(questions)} ---")
        safe_print(f"[{env_name}] ❓ 输入问题: {question_text if question_text else '【无文本，仅上传文件】'}")
        safe_print(
            f"[{env_name}] 🤖 目标Agent: {target_agent if target_agent and str(target_agent).strip() else '【未指定】'}")
        safe_print(f"[{env_name}] ⚙️ 最终执行状态: {CURRENT_STATE}")

        try:
            if is_tc_ui:
                safe_print(f"[{env_name}] 🌐 [语言切换]: 正在切换为 繁中...")
                lang_container = WebDriverWait(driver, 3).until(
                    EC.presence_of_element_located((By.CLASS_NAME, "language-container"))
                )
                driver.execute_script("arguments[0].click();", lang_container)
                time.sleep(1)
                tc_option = WebDriverWait(driver, 3).until(
                    EC.presence_of_element_located(
                        (By.XPATH, "//a[contains(@class, 'language-option') and text()='繁中']"))
                )
                driver.execute_script("arguments[0].click();", tc_option)
                time.sleep(1.5)
            else:
                safe_print(f"[{env_name}] 🌐 [语言切换]: 正在切换为 英文 (默认)...")
                lang_container = WebDriverWait(driver, 3).until(
                    EC.presence_of_element_located((By.CLASS_NAME, "language-container"))
                )
                driver.execute_script("arguments[0].click();", lang_container)
                time.sleep(1)
                eng_option = WebDriverWait(driver, 3).until(
                    EC.presence_of_element_located(
                        (By.XPATH, "//a[contains(@class, 'language-option') and text()='ENG']"))
                )
                driver.execute_script("arguments[0].click();", eng_option)
                time.sleep(1)

            if not is_text_only and file_path:
                file_input = WebDriverWait(driver, 10).until(
                    EC.presence_of_element_located((By.XPATH, "//input[@type='file']"))
                )
                file_input.send_keys(file_path)
                time.sleep(1)
            else:
                safe_print(f"[{env_name}] 📄 [跳过]: 纯文本模式，无需上传文件。")

            # ================= 问题输入 (JS 强行注入版本对齐) =================
            try:
                text_area = WebDriverWait(driver, 10).until(
                    EC.presence_of_element_located((By.ID, "background-info"))
                )
                driver.execute_script("arguments[0].value = '';", text_area)
                driver.execute_script("arguments[0].value = arguments[1];", text_area, question_text)
                driver.execute_script("""
                    var input = arguments[0];
                    var lastValue = input.value;
                    input.value = arguments[1];
                    var event = new Event('input', { bubbles: true });
                    // 破解 React 16+ 内部的值追踪器
                    var tracker = input._valueTracker;
                    if (tracker) { tracker.setValue(lastValue); }
                    input.dispatchEvent(event);
                """, text_area, question_text)
                driver.execute_script("arguments[0].dispatchEvent(new Event('change', { bubbles: true }));", text_area)
                safe_print(f"[{env_name}] ✅ [JS 注入成功]: 已成功在后台写入问题文本。")
                time.sleep(1)
            except Exception as input_err:
                safe_print(f"[{env_name}] ❌ 输入问题文本时发生异常: {input_err}")
                raise input_err

            safe_print(f"[{env_name}] 🔍 [通用步骤]: 正在点击设置齿轮图标...")
            gear_btn = WebDriverWait(driver, 5).until(
                EC.presence_of_element_located((By.XPATH, "//i[contains(@class, 'bi-gear-fill')]"))
            )
            driver.execute_script("arguments[0].click();", gear_btn)
            time.sleep(0.5)

            if search_candidates:
                # 情况 A：Excel 指定了 Agent，去页面上找
                agent_success = False
                last_err = None
                for candidate in search_candidates:
                    try:
                        safe_print(f"[{env_name}] 👉 正在尝试选中 Agent: '{candidate}' ...")
                        execute_state(driver, CURRENT_STATE, candidate)
                        agent_success = True
                        safe_print(f"[{env_name}] ✅ 成功找到并应用 Agent: '{candidate}'")
                        break
                    except Exception as e:
                        safe_print(f"[{env_name}] ⚠️ 网站未找到 '{candidate}'，准备尝试下一个...")
                        last_err = e

                if not agent_success:
                    safe_print(
                        f"[{env_name}] ⚠️ [保底机制触发]: 候选 Agent {search_candidates} 均无法找到！自动降级为 State 2。")
                    CURRENT_STATE = "2"
                    target_agent = str(target_agent) + " (未找到，降级为 State 2)"
                    execute_state(driver, CURRENT_STATE)
            else:
                # 情况 B：Excel 原本就没有指定 Agent，名正言顺地直接执行通用模式
                safe_print(f"[{env_name}] ℹ️ 目标 Agent 为空，直接应用通用模式 (State 2)...")
                execute_state(driver, CURRENT_STATE)

            safe_print(f"[{env_name}] 🔍 [通用步骤]: 面板配置完毕，正在点击 Apply Settings...")
            apply_btn = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.ID, "global-settings-apply-btn"))
            )
            driver.execute_script("arguments[0].click();", apply_btn)
            time.sleep(1)

            safe_print(f"[{env_name}] 🚀 [通用步骤]: 正在点击发送/生成按钮...")
            submit_btn = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.XPATH, "//button[contains(@class, 'btn-enter')]"))
            )
            driver.execute_script("arguments[0].click();", submit_btn)

            handle_post_send(driver, CURRENT_STATE)
            safe_print(f"[{env_name}] ✅ 已成功进入自动生成等待环节...")

            # ================= 智能监控 AI 生成进度 (对齐 main.py) =================
            safe_print(f"[{env_name}] ⏳ 正在智能监控 AI 生成进度...")
            try:
                time.sleep(5)
                last_length = 0
                stable_count = 0
                max_wait_loops = 300
                required_stable_seconds = 5

                for _ in range(max_wait_loops):
                    if STOP_SCRIPT:
                        safe_print(f"[{env_name}] 🛑 收到中止指令，立即停止网页监控！")
                        break
                    try:
                        _ = driver.window_handles
                    except Exception:
                        safe_print(f"\n[{env_name}] 🚨 监控期间侦测到浏览器窗口被关闭，退出！")
                        raise Exception("window already closed")
                    time.sleep(1)

                    current_length = 0
                    try:
                        previews = driver.find_elements(By.ID, "preview")
                        for p in reversed(previews):
                            if p.is_displayed():
                                actual_text = p.get_attribute("innerText") or p.get_attribute("textContent") or ""
                                current_length = len(actual_text.strip())
                                break
                    except Exception:
                        pass

                    if current_length > 0 and current_length == last_length:
                        stable_count += 1
                    else:
                        stable_count = 0

                    last_length = current_length

                    if stable_count >= required_stable_seconds:
                        safe_print(
                            f"[{env_name}] ✅ 回答文本连续 {required_stable_seconds} 秒无变化，彻底完成！字数: {current_length}")
                        break
                else:
                    if not STOP_SCRIPT:
                        safe_print(f"[{env_name}] ⚠️ 警告：监控达到 300 秒上限，生成总时间超时！")
                        timeout_status = "yes (总时间超时)"
            except Exception as wait_error:
                safe_print(f"[{env_name}] ⚠️ 智能监控发生异常，强制继续执行: {wait_error}")
                time.sleep(2)

            if STOP_SCRIPT: break

            current_page_url = driver.current_url
            safe_print(f"[{env_name}] 🔗 已抓取 shared link: {current_page_url}")

            # ================= 提取网页回答 (对齐 main.py) =================
            safe_print(f"[{env_name}] 📥 正在提取生成的回答...")
            try:
                safe_print(f"[{env_name}] ⏳ 正在等待并提取最新的可见回答 (最多等待 60 秒)...")

                def get_valid_preview(d):
                    try:
                        _ = d.window_handles
                    except Exception:
                        return "BROWSER_CLOSED"
                    previews = d.find_elements(By.ID, "preview")
                    for p in reversed(previews):
                        actual_content = p.get_attribute("innerText") or p.get_attribute("textContent") or ""
                        if len(actual_content.strip()) > 0:
                            return actual_content
                    return False

                # 替换为可立刻响应 Ctrl+C 的秒级循环：
                answer_text = ""
                wait_time = 0
                max_wait_time = 60

                while wait_time < max_wait_time:
                    if STOP_SCRIPT:
                        safe_print(f"[{env_name}] 🛑 收到中止指令，立即中断提取动作！")
                        break
                    try:
                        res = get_valid_preview(driver)
                        if res == "BROWSER_CLOSED":
                            safe_print(f"[{env_name}] 🚨 提取内容时侦测到浏览器已关闭，终止当前任务！")
                            break
                        if res:  # 如果成功抓到了文字
                            answer_text = res
                            break
                    except Exception:
                        pass

                    time.sleep(1)
                    wait_time += 1

                if answer_text == "BROWSER_CLOSED":
                    break  # 直接跳出当前环境的主循环

                if not answer_text or not answer_text.strip():
                    safe_print(f"[{env_name}] ⚠️ 等待 60 秒后未抓取到内容！")
                    answer_text = "提取文本失败/为空"
                    if timeout_status == "No":
                        timeout_status = "yes (生成超时)"
                    else:
                        timeout_status += " & yes (生成超时)"
                else:
                    safe_print(f"[{env_name}] ✅ 成功提取到回答，长度: {len(answer_text)} 字符")

                prep_time = "N/A"
                comp_time = "N/A"
                try:
                    # 1. 强制等待 3 秒，等组件完全挂载到 React 树上
                    time.sleep(3)

                    # 2. 注入核弹级 JS：全方位透视 DOM 与 React 底层内存树
                    times_info = driver.execute_script("""
                                        try {
                                            var prep = 'N/A', comp = 'N/A';
                                            var debugLog = [];
                                            // 锁定我们要找的滚动组件标签
                                            var flows = document.querySelectorAll('number-flow-react, number-flow');
                                            debugLog.push('找到标签数: ' + flows.length);
                                            var vals = [];

                                            for(var i=0; i<flows.length; i++) {
                                                var el = flows[i];
                                                var val = null;

                                                // 第一层：尝试常规 DOM 属性
                                                if (el.value !== undefined) val = el.value;
                                                else if (el.getAttribute('value')) val = el.getAttribute('value');
                                                else if (el.getAttribute('aria-valuenow')) val = el.getAttribute('aria-valuenow');

                                                // 第二层：入侵 React 16-18+ 的 Fiber 内存树，向上回溯寻找 memoizedProps
                                                if (val == null) {
                                                    var fiberKey = Object.keys(el).find(k => k.startsWith('__reactFiber$'));
                                                    if (fiberKey) {
                                                        var curr = el[fiberKey];
                                                        // 向上回溯 5 层，寻找包含 value 变量的节点
                                                        for (var j=0; j<5; j++) {
                                                            if (curr && curr.memoizedProps && curr.memoizedProps.value !== undefined) {
                                                                val = curr.memoizedProps.value;
                                                                debugLog.push('通过Fiber提取成功');
                                                                break;
                                                            }
                                                            if (curr) curr = curr.return;
                                                        }
                                                    }
                                                }

                                                // 第三层：入侵 React 17+ 的 Props 内存
                                                if (val == null) {
                                                    var propsKey = Object.keys(el).find(k => k.startsWith('__reactProps$'));
                                                    if (propsKey && el[propsKey] && el[propsKey].value !== undefined) {
                                                        val = el[propsKey].value;
                                                        debugLog.push('通过Props提取成功');
                                                    }
                                                }

                                                if (val !== null) vals.push(val);
                                            }

                                            if (vals.length >= 1) prep = vals[0] + 's';
                                            if (vals.length >= 2) comp = vals[1] + 's';

                                            return { prep: prep, comp: comp, debug: debugLog.join(' | ') };
                                        } catch(err) {
                                            return { prep: 'N/A', comp: 'N/A', debug: 'JS报错: ' + err.message };
                                        }
                                    """)

                    if times_info:
                        prep_time = times_info.get("prep", "N/A")
                        comp_time = times_info.get("comp", "N/A")
                        debug_info = times_info.get("debug", "")

                        # 第四层极限保底：如果连 React 内存里都没有，强行用正则刮取底层 HTML 代码
                        if prep_time == "N/A":
                            html = driver.execute_script("return document.body.innerHTML;")
                            # 匹配类似 value="99.5" 的隐藏属性
                            html_matches = re.findall(
                                r'<number-flow[^>]*?(?:value|aria-valuenow)=["\']?([0-9.]+)["\']?', html,
                                re.IGNORECASE)
                            if len(html_matches) >= 1:
                                prep_time = html_matches[0] + "s"
                                debug_info += " | 触发底层HTML正则兜底"
                            if len(html_matches) >= 2:
                                comp_time = html_matches[1] + "s"

                        safe_print(
                            f"[{i + 1}] ⏱️ 提取时间 -> 准备耗时: {prep_time}, 完成耗时: {comp_time} (底层诊断: {debug_info})")
                    else:
                        safe_print(f"[{i + 1}] ⏱️ 提取失败：JS 脚本未返回任何数据。")

                except Exception as time_err:
                    safe_print(f"[{i + 1}] ⚠️ 提取时间发生代码异常: {time_err}")

                if is_text_only:
                    file_content = "【无原始文档，用户仅提供了纯文本提问，请仅根据问题本身评估回答是否准确且符合逻辑】"
                else:
                    file_content = read_file_content(file_path, env_name)

                if "yes" in timeout_status.lower():
                    safe_print(f"[{env_name}] ⚠️ 侦测到超时，跳过 DeepSeek 评价...")
                    eval_results = {
                        "tester_expectation": "Failed",
                        "input_language": "N/A",
                        "output_language": "N/A",
                        "language_status": "Failed",
                        "evaluation": "生成超时或提取失败，未能获取有效回答。",
                        "reference_link": "N/A",
                        "document_contain_citations": "None"
                    }
                else:
                    safe_print(f"[{env_name}] 🤖 等待 DeepSeek 进行语言检测与质量评价...")
                    eval_results = evaluate_with_deepseek(question_text, file_content, answer_text, target_language)

                tester_exp = eval_results.get("tester_expectation", "Unknown")
                input_lang = eval_results.get("input_language", "Unknown")
                output_lang = eval_results.get("output_language", "Unknown")
                lang_status = eval_results.get("language_status", "Unknown")
                evaluation_text = eval_results.get("evaluation", "Unknown")
                ref_link = eval_results.get("reference_link", "N/A")
                doc_contain = eval_results.get("document_contain_citations", "None")

                try:
                    wb = openpyxl.load_workbook(excel_path)
                    ws = wb.active
                    short_answer = answer_text[:100] + "..." if len(answer_text) > 100 else answer_text
                    display_filename = filename if filename else ""
                    display_target_language = target_language if target_language else "N/A"

                    ws.append([
                        i + 1,  # label
                        question_text,  # Request
                        display_filename,  # filename
                        "No",  # Crash
                        tester_exp,  # Tester Expectation
                        display_target_language, input_lang, output_lang,
                        lang_status, short_answer, current_page_url, evaluation_text,
                        target_agent if target_agent else "未指定", ref_link,
                        doc_contain, prep_time, comp_time, timeout_status
                    ])
                    wb.save(excel_path)
                    safe_print(f"[{env_name}] ✅ 评价结果已成功写入 Excel 文件。")
                except Exception as excel_err:
                    safe_print(f"[{env_name}] ❌ 写入 Excel 时失败: {excel_err}")

            except Exception as extract_error:
                safe_print(f"[{env_name}] ❌ 提取文本、评价或保存数据时发生错误: {extract_error}")

            try:
                safe_print(f"[{env_name}] 🌐 [语言重置]: 正在准备返回前重置为英文...")
                globe_icon = WebDriverWait(driver, 3).until(
                    EC.presence_of_element_located((By.CLASS_NAME, "language-icon"))
                )
                driver.execute_script("arguments[0].click();", globe_icon)
                time.sleep(1)
                eng_reset_option = WebDriverWait(driver, 3).until(
                    EC.presence_of_element_located(
                        (By.XPATH, "//a[contains(@class, 'language-option') and text()='ENG']"))
                )
                driver.execute_script("arguments[0].click();", eng_reset_option)
            except Exception as reset_lang_err:
                safe_print(f"[{env_name}] ⚠️ 重置语言步骤失败，继续返回首页: {reset_lang_err}")

            try:
                safe_print(f"[{env_name}] 🏠 正在尝试通过点击 Logo 返回首页...")
                home_btn = WebDriverWait(driver, 7).until(
                    EC.element_to_be_clickable((By.CLASS_NAME, "image-home"))
                )
                driver.execute_script("arguments[0].click();", home_btn)
                safe_print(f"[{env_name}] ✅ 成功通过点击 Logo 返回")
            except Exception:
                safe_print(f"[{env_name}] ⚠️ 未识别到主页 Logo，直接重新请求首页网址...")
                # 尝试清空 SessionStorage（清理上一个对话可能留下的脏数据）
                try:
                    driver.execute_script("window.sessionStorage.clear();")
                except:
                    pass
                driver.get(home_url)

            time.sleep(1)
            close_popups(driver, env_name)
            if (i + 1) % 50 == 0:
                safe_print(f"[{env_name}] [{i + 1}] 🔄 触发内存保护：该环境已连续运行 50 次，正在重启释放内存...")
                try:
                    driver.quit()
                except:
                    pass

                # 1. 重新拉起干净的浏览器实例
                driver = webdriver.Chrome(service=Service(chrome_driver_path), options=chrome_options)
                driver.get(home_url)
                time.sleep(2)

                # 2. 重新执行自动登录逻辑
                safe_print(f"[{env_name}] [{i + 1}] 🔑 正在为新浏览器重新执行登录...")
                try:
                    login_nav = WebDriverWait(driver, 10).until(
                        EC.element_to_be_clickable((By.ID, "nav-login"))
                    )
                    driver.execute_script("arguments[0].click();", login_nav)

                    username_input = WebDriverWait(driver, 10).until(
                        EC.presence_of_element_located((By.ID, "sso-username"))
                    )
                    password_input = driver.find_element(By.ID, "sso-password")

                    driver.execute_script("arguments[0].value = arguments[1];", username_input, username)
                    driver.execute_script("arguments[0].dispatchEvent(new Event('input', { bubbles: true }));",
                                          username_input)
                    driver.execute_script("arguments[0].value = arguments[1];", password_input, password)
                    driver.execute_script("arguments[0].dispatchEvent(new Event('input', { bubbles: true }));",
                                          password_input)

                    submit_btn = driver.find_element(By.CLASS_NAME, "sso-submit-btn")
                    driver.execute_script("arguments[0].click();", submit_btn)

                    close_popups(driver, env_name)
                    safe_print(f"[{env_name}] [{i + 1}] ✅ 内存释放完毕，登录成功，继续后续任务...")
                    time.sleep(1)
                    close_popups(driver, env_name)
                except Exception as re_login_err:
                    safe_print(f"[{env_name}] [{i + 1}] ❌ 重启浏览器后自动登录失败: {re_login_err}，该环境线程终止。")
                    break

        except Exception as e:
            # 静默过滤掉手动关闭窗口的冗长报错，保持控制台整洁
            error_msg = str(e).lower()
            is_manually_closed = any(k in error_msg for k in [
                "window already closed", "target window already closed",
                "no such window", "disconnected", "connection refused",
                "not connected to devtools", "aborted", "invalid session id"
            ])
            if is_manually_closed:
                safe_print(f"[{env_name}] 🚨 侦测到浏览器窗口已被手动关闭，跳出任务循环！")
                break

            safe_print(f"[{env_name}] ❌ [报错] 处理 {filename if filename else '纯文本'} 时发生错误: {e}")

            # 📝 兜底写入 Excel 逻辑，防止数据漏行
            safe_print(f"[{env_name}] 📝 正在将崩溃记录写入 Excel...")
            try:
                wb = openpyxl.load_workbook(excel_path)
                ws = wb.active

                display_filename = filename if filename else ""
                display_target_language = target_language if target_language else "N/A"
                safe_agent = target_agent if target_agent else "未指定"
                crash_reason = f"Automation Crash: {str(e)[:150]}"
                error_info = f"⚠️ 自动化执行崩溃: {str(e)[:150]}"

                # 严格按照表头顺序追加“尸体”记录
                ws.append([
                    i + 1, question_text, display_filename, crash_reason,
                    "Failed", display_target_language, "N/A", "N/A",
                    "Failed", "【执行崩溃，未能生成回答】", "N/A", error_info,
                    safe_agent, "N/A", "None", "N/A", "N/A", "Crash (Error)"
                ])
                wb.save(excel_path)
                safe_print(f"[{env_name}] ✅ 崩溃记录已成功保存，数据未流失。")
            except Exception as excel_err:
                safe_print(f"[{env_name}] ❌ 紧急写入兜底记录至 Excel 时失败: {excel_err}")

            # 现场清理与重试
            safe_print(f"[{env_name}] 🔄 正在清理现场并重新加载首页...")
            try:
                driver.get(home_url)
                time.sleep(2)
            except Exception:
                pass

    safe_print(f"[{env_name}] 📊 正在生成最终的 Summary 报告...")
    try:
        dynamic_csv_name = f"Summary_{env_name}_{base_testcase_name}_{timestamp}.csv"
        summary_dir = os.path.join(project_dir, f"Summaries_{env_name}")
        os.makedirs(summary_dir, exist_ok=True)
        output_csv = os.path.join(summary_dir, dynamic_csv_name)
        generate_summary_csv(excel_path, output_csv)
        safe_print(f"[{env_name}] ✅ 汇总报告已生成: {output_csv}")
    except Exception as e:
        safe_print(f"[{env_name}] ⚠️ 生成汇总报告时发生异常: {e}")
    if STOP_SCRIPT:
        safe_print(f"[{env_name}] \n🛑 任务已被手动中断！")
    else:
        safe_print(f"[{env_name}] \n✅ 该环境所有测试用例已全部运行完毕！")


if __name__ == "__main__":
    # 1. 预加载浏览器驱动
    chrome_driver_path = ChromeDriverManager().install()

    listener_thread = threading.Thread(target=listen_for_hotkey, daemon=True)
    listener_thread.start()

    print("🚀 正在初始化全局资源 (读取 Excel)...")
    # 2. 【关键修复】主线程必须在这里读取 Excel 数据，否则 test_data 变量不存在
    test_data = get_test_data_from_excel(INPUT_EXCEL_PATH)

    print("🚀 开始多线程并行执行自动化任务...")

    # 3. 【关键修复】必须把 test_data 和 chrome_driver_path 塞进 args 元组里传给线程，并加上 daemon=True
    t1 = threading.Thread(
        target=run_automation,
        args=(URL_1, ENV_NAME_1, USERNAME_1, PASSWORD_1, test_data, chrome_driver_path),
        daemon=True
    )
    t2 = threading.Thread(
        target=run_automation,
        args=(URL_2, ENV_NAME_2, USERNAME_2, PASSWORD_2, test_data, chrome_driver_path),
        daemon=True
    )

    t1.start()
    t2.start()

    while t1.is_alive() or t2.is_alive():
        time.sleep(1)
        if STOP_SCRIPT:
            break

    print("👉 所有自动化任务均已结束。控制权已释放（浏览器将由 detach 属性保持开启）。")
    while True:
        if STOP_SCRIPT:
            print("👋 侦测到 Ctrl+C 退出指令，控制台自动关闭！")
            keyboard.unhook_all()
            break
        # 注意：这里不需要检测 driver，因为主线程不持有 driver，只需阻塞住控制台即可
        time.sleep(1)
