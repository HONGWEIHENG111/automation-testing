import os
import time
import queue
import re
from datetime import datetime
import threading
import openpyxl
import keyboard
from webdriver_manager.chrome import ChromeDriverManager
from utils.logger import setup_global_logger
from utils.config import (
    INPUT_EXCEL_PATH, MAX_WORKERS,
    URL_1, ENV_NAME_1, USERNAME_1, PASSWORD_1,
    URL_2, ENV_NAME_2, USERNAME_2, PASSWORD_2
)
from utils.common import (
    show_authorship, close_popups, get_test_data_from_excel, init_browser, perform_login
)
from tools.summary_generator import generate_summary_csv

# ================= 引入 Core 模块 =================
from utils.signals import STOP_EVENT
from utils.models import TaskInput
from utils.pipeline import process_single_task

PRINT_LOCK = threading.Lock()


def safe_print(*args, **kwargs):
    with PRINT_LOCK:
        print(*args, **kwargs)


def listen_for_hotkey():
    try:
        keyboard.add_hotkey('ctrl+alt+h', show_authorship)
        keyboard.wait('ctrl+q')
        STOP_EVENT.set()
        print("\n\n🛑 [紧急刹车] 侦测到 Ctrl+Q 键按下！脚本将在当前步骤跳出...\n")
    except ImportError:
        print("⚠️ 缺少 keyboard 模块，无法使用快捷键终止功能。")
    except Exception as e:
        print(f"⚠️ 热键监听启动失败 (可能缺少管理员权限)，将禁用 Ctrl+Q 功能: {e}")

def excel_writer_worker(excel_path, q, log_func=print):
    """专职 Excel 写入后台线程（带非法字符过滤）"""
    import openpyxl
    try:
        wb = openpyxl.load_workbook(excel_path)
        ws = wb.active
    except Exception as e:
        log_func(f"❌ [写线程致命错误] 无法加载 Excel: {e}")
        return

    while True:
        item = q.get()
        if item == "STOP":
            try: wb.save(excel_path)
            except: pass
            break

        task, result = item
        try:
            row_data = [
                task.index + 1, task.question_text, task.filename if task.filename else "",
                result.crash_reason, result.tester_expectation, task.target_language if task.target_language else "N/A",
                result.input_language, result.output_language, result.language_status,
                result.answer_text, result.shared_link, result.evaluation_text,
                result.actual_agent_used, result.reference_link, result.document_contain,
                result.prep_time, result.comp_time, result.timeout_status
            ]
            target_row = task.index + 2
            for col_index, value in enumerate(row_data, start=1):
                # 🛡️ 核心防御：过滤 Excel 不支持的非法控制字符
                if isinstance(value, str):
                    value = re.sub(r'[\x00-\x08\x0b\x0c\x0e-\x1f]', '', value)
                ws.cell(row=target_row, column=col_index, value=value)

            save_success = False
            for retry in range(3):
                try:
                    wb.save(excel_path)
                    save_success = True
                    break
                except PermissionError:
                    time.sleep(3)
            if not save_success:
                log_func(f"❌ [致命错误] 多次尝试保存失败，第 {task.index + 1} 行结果丢失！")
        except Exception as e:
            log_func(f"❌ [写线程异常] 写入行 {task.index + 1} 失败: {e}")
        finally:
            q.task_done()

def run_automation_env(env_name, home_url, username, password, task_inputs, test_dir, project_dir, cached_driver_path):
    """
    环境独立调度器：管理特定环境的登录、Excel 创建和任务循环。
    """
    custom_log = lambda msg: safe_print(f"[{env_name}] {msg}")

    # ================= 初始化专属 Excel =================
    base_testcase_name = os.path.splitext(os.path.basename(INPUT_EXCEL_PATH))[0]
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    dynamic_filename = f"evaluation_results_{env_name}_{base_testcase_name}_{timestamp}.xlsx"
    excel_dir = os.path.join(project_dir, f"Evaluation_Results_{env_name}")
    os.makedirs(excel_dir, exist_ok=True)
    excel_path = os.path.join(excel_dir, dynamic_filename)

    if not os.path.exists(excel_path):
        wb = openpyxl.Workbook()
        wb.properties.creator = "Henry HONG "
        wb.properties.description = "Authored by Henry HONG. "
        ws = wb.active
        ws.title = "Evaluation Results"
        ws.append(["label", "Request", "filename", "Crash", "Tester Expectation", "Selected Language", "Input Language",
                   "Output Language", "Language Overall Status", "answer", "shared link", "DeepSeek评价内容",
                   "Selected agent", "Reference Link", "Document Contain[1][2][3]", "Preparation Time",
                   "Completion Time", "Timeout_States"])
        wb.save(excel_path)
        custom_log(f"📊 已创建评价结果 Excel 文件: {excel_path}")

    result_queue = queue.Queue()
    writer_thread = threading.Thread(target=excel_writer_worker, args=(excel_path, result_queue, custom_log),
                                     daemon=True)
    writer_thread.start()

    # ================= 启动环境专属浏览器 =================
    driver = None
    try:
        driver = init_browser(log_func=custom_log, chrome_driver_path=cached_driver_path)
        driver.get(home_url)
        time.sleep(1)
        perform_login(driver, username, password, log_func=custom_log)
    except Exception as e:
        custom_log(f"❌ 浏览器启动失败: {e}")
        try:
            if driver:
                driver.quit()  # 账号密码错了或者崩溃，自动把浏览器关掉
        except:
            pass
        return

    # ================= 顺序循环执行当前环境的任务 =================
    try:
        for task in task_inputs:
            if STOP_EVENT.is_set():
                custom_log("🛑 收到中止指令，停止后续任务。")
                break

            try:
                from selenium.webdriver.support.ui import WebDriverWait
                from selenium.webdriver.support import expected_conditions as EC
                WebDriverWait(driver, 0.5).until(EC.alert_is_present())
                driver.switch_to.alert.accept()
            except:
                pass

            try:
                driver.get(home_url)
                time.sleep(1.5)
            except:
                pass

            custom_log(f"\n--- 进度: {task.index + 1}/{len(task_inputs)} ---")
            status = process_single_task(
                task=task,
                driver=driver,
                test_dir=test_dir,
                result_queue=result_queue,
                stop_event=STOP_EVENT,
                log_func=custom_log
            )

            if status == "BROWSER_CLOSED": break

            # 环境内的 50 次内存保护
            if (task.index + 1) % 50 == 0 and not STOP_EVENT.is_set():
                custom_log(f"[{task.index + 1}] 🔄 触发内存保护：连续运行 50 次，重启释放内存...")
                try:
                    driver.quit()
                except:
                    pass

                try:
                    driver = init_browser(log_func=custom_log, chrome_driver_path=cached_driver_path)
                    driver.get(home_url)
                    time.sleep(2)
                    perform_login(driver, username, password, log_func=custom_log)
                    custom_log(f"[{task.index + 1}] ✅ 内存释放完毕，新浏览器登录成功...")
                except Exception as e:
                    custom_log(f"[{task.index + 1}] ❌ 重启浏览器失败: {e}，该环境终止。")
                    break
    except Exception as e:  # <--- 新增
        custom_log(f"❌ 环境发生致命错误: {e}")
    finally:
    # ================= 生成当前环境的报告 =================
        custom_log("💾 正在等待当前环境的残余数据写入 Excel...")
        if 'result_queue' in locals():
            result_queue.put("STOP")
            writer_thread.join(timeout=10)
        custom_log("📊 正在生成最终的 Summary 报告...")
        from utils.config import SUMMARY_ENV_1_DIR, SUMMARY_ENV_2_DIR
        target_dir = SUMMARY_ENV_1_DIR if env_name == ENV_NAME_1 else SUMMARY_ENV_2_DIR
        output_csv = os.path.join(target_dir, f"Summary_{env_name}_{base_testcase_name}_{timestamp}.csv")
        os.makedirs(target_dir, exist_ok=True)

        try:
            generate_summary_csv(excel_path, output_csv)
            custom_log(f"✅ 汇总报告已生成: {output_csv}")
        except Exception as e:
            custom_log(f"⚠️ 生成汇总报告时发生异常: {e}")

        if not STOP_EVENT.is_set():
            custom_log("\n✅ 该环境所有测试用例运行完毕！")


def run_automation():
    setup_global_logger()
    print("📥 正在初始化並緩存瀏覽器驅動...")
    cached_driver_path = ChromeDriverManager().install()
    threading.Thread(target=listen_for_hotkey, daemon=True).start()

    project_dir = os.path.dirname(os.path.abspath(__file__))
    test_dir = os.path.join(project_dir, "test")
    if not os.path.exists(test_dir):
        print(f"⚠️ [警告] 找不到测试文件夹: {test_dir}，已自动创建。若本次测试包含文件上传，请确保文件放入该目录。")
        os.makedirs(test_dir, exist_ok=True)
    questions, selected_agents, target_filenames, selected_languages = get_test_data_from_excel(INPUT_EXCEL_PATH)
    # ================= 文件数量校验 =================
    valid_extensions = ('.pdf', '.docx', '.csv', '.txt')
    test_files = [f for f in os.listdir(test_dir) if f.lower().endswith(valid_extensions)]
    expected_file_count = len([f for f in target_filenames if str(f).strip()])
    if len(test_files) != expected_file_count:
        print(f"⚠️ 文件夹中的文件数量 ({len(test_files)}) 与 Excel 中的问题数量 ({len(questions)}) 不匹配！")
    task_inputs = []
    for i in range(len(questions)):
        task_inputs.append(TaskInput(
            index=i,
            question_text=questions[i],
            target_agent=selected_agents[i],
            filename=target_filenames[i],
            target_language=selected_languages[i]
        ))

    print("🚀 开始多环境并行执行自动化任务...")

    env_configs = [
        (ENV_NAME_1, URL_1, USERNAME_1, PASSWORD_1),
        (ENV_NAME_2, URL_2, USERNAME_2, PASSWORD_2)
    ]
    active_envs = env_configs[:MAX_WORKERS]
    threads = []

    for env_args in active_envs:
        t = threading.Thread(
            target=run_automation_env,
            args=(*env_args, task_inputs, test_dir, project_dir, cached_driver_path),
            daemon=True
        )
        t.start()
        threads.append(t)
    try:
        while any(t.is_alive() for t in threads):
            time.sleep(1)
    except KeyboardInterrupt:
        print("\n🛑 侦测到系统强行中断信号 (Ctrl+Q)！正在通知所有环境紧急保存并生成报告...")
        STOP_EVENT.set()
        # 必须等待子线程把 finally 里的报告生成完毕，主线程才能死
        for t in threads:
            t.join(timeout=30)
    if STOP_EVENT.is_set():
        print("\n🛑 侦测到任务被手动中断，所有环境的善后报告已生成完毕。")
    else:
        print("\n👉 侦测到所有环境自动化任务已完美结束。")

    print("👋 脚本即将安全退出，保留浏览器现场供排查！")


if __name__ == "__main__":
    run_automation()
