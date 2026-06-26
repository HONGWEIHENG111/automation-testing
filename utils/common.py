import os
import time
import base64
import difflib
import openpyxl
import PyPDF2
from docx import Document

from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.edge.service import Service as EdgeService
from selenium.webdriver.edge.options import Options as EdgeOptions
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
from webdriver_manager.microsoft import EdgeChromiumDriverManager


def show_authorship():
    secret_msg = base64.b64decode(b'Q29kZSBBdXRob3I6IEhlbnJ5IEhPTkc=').decode('utf-8')
    print(f"\n\n====================================")
    print(f"🌟 {secret_msg} 🌟")
    print(f"====================================\n\n")


def close_popups(driver, log_func=print):
    """專門用於檢測並關閉頁面上的干擾彈窗"""
    log_func("🔍 正在执行弹窗清理...")
    try:
        skip_btn = WebDriverWait(driver, 1).until(
            EC.presence_of_element_located((By.ID, "gen-tour-welcome-skip"))
        )
        driver.execute_script("arguments[0].click();", skip_btn)
        log_func("✅ 已成功跳过 Welcome 导览弹窗")
        time.sleep(0.5)
    except Exception:
        pass

    try:
        dont_show_label = WebDriverWait(driver, 1).until(
            EC.presence_of_element_located((By.XPATH, "//label[@for='dontShowAgain']"))
        )
        driver.execute_script("arguments[0].click();", dont_show_label)
        log_func("✅ 已勾选 'Don't show this again'")
        time.sleep(0.5)

        close_btn = driver.find_element(By.CLASS_NAME, "credit-popup-close")
        driver.execute_script("arguments[0].click();", close_btn)
        log_func("✅ 已成功关闭 News 弹窗")
    except Exception:
        pass


def get_test_data_from_excel(excel_path, log_func=print):
    """從 Excel 中智能尋找表頭並提取資料"""
    questions, selected_agents, filenames, selected_languages = [], [], [], []

    if not os.path.exists(excel_path):
        log_func(f"[致命错误] 找不到 Excel 文件: {excel_path}")
        return questions, selected_agents, filenames, selected_languages

    try:
        wb = openpyxl.load_workbook(excel_path)
        ws = wb.active
        request_col = agent_col = filename_col = language_col = header_row = None

        for row in ws.iter_rows():
            # 【严谨机制 1】：每次扫描新的一行时，准备一套干净的临时变量
            temp_req = temp_agent = temp_file = temp_lang = None
            for cell in row:
                if cell.value:
                    val = str(cell.value).strip().lower()
                    if val == "request":
                        temp_req = cell.column
                    elif val == "selected agent":
                        temp_agent = cell.column
                    elif val in ["filename", "file name", "file_name"]:
                        temp_file = cell.column
                    else:
                        lang_keywords = ["selected language", "selected_language", "language", "lang", "目標語言",
                                       "测试语言", "ui语言", "ui language"]
                        is_language_column = False
                        if val in lang_keywords or "language" in val or "lang" in val or "語言" in val:
                            is_language_column = True
                        else:
                            matches = difflib.get_close_matches(val, lang_keywords, n=1, cutoff=0.6)
                            if matches: is_language_column = True
                        if is_language_column:
                            temp_lang = cell.column

            # 【严谨机制 2】：等这一整行的所有列都老老实实扫完后，再做终极判定
            # 只要发现了 Request 或 Filename 任意一个核心锚点，就确认这一行确实是表头
            if temp_req is not None or temp_file is not None:
                header_row = row[0].row
                request_col = temp_req
                agent_col = temp_agent
                filename_col = temp_file
                language_col = temp_lang
                break  # 确认找到表头后，安全跳出外层行循环，不再扫描下方的数据行

        if not request_col and not filename_col:
            log_func(
                f"[致命错误] 在 Excel({excel_path}) 中找不到 'Request' 也找不到 'Filename' 表頭！無法提供任何測試輸入。")
            return questions, selected_agents, filenames, selected_languages

        if not filename_col: log_func(f"[降級提示] 找不到 'Filename' 表頭，將自動降級為【純文本模式】運行。")
        if not agent_col: log_func(f"[降級提示] 找不到 'Selected Agent' 表頭，將自動降級為【狀態 2 (通用自動模式)】。")
        if not language_col: log_func(f"[降級提示] 找不到 'Language' 表頭，將自動使用【英文 (默認)】。")

        for row in range(header_row + 1, ws.max_row + 1):
            req_val = ws.cell(row=row, column=request_col).value if request_col else ""
            file_val = ws.cell(row=row, column=filename_col).value if filename_col else ""
            req_str, file_str = str(req_val).strip() if req_val else "", str(file_val).strip() if file_val else ""

            if req_str or file_str:
                questions.append(req_str)
                filenames.append(file_str)
                selected_agents.append(
                    str(ws.cell(row=row, column=agent_col).value).strip() if agent_col and ws.cell(row=row,
                                                                                                   column=agent_col).value else "")
                selected_languages.append(
                    str(ws.cell(row=row, column=language_col).value).strip() if language_col and ws.cell(row=row,
                                                                                                         column=language_col).value else "")

    except Exception as e:
        log_func(f"[格式錯誤] 讀取 Excel 文件失敗: {e}")

    return questions, selected_agents, filenames, selected_languages


def read_file_content(file_path, log_func=print):
    """根據文件後綴讀取不同格式的文件內容"""
    ext = os.path.splitext(file_path)[1].lower()
    content = ""
    try:
        if ext in ['.txt', '.csv']:
            try:
                with open(file_path, 'r', encoding='utf-8') as f:
                    content = f.read()
            except UnicodeDecodeError:
                with open(file_path, 'r', encoding='gbk') as f:
                    content = f.read()
        elif ext == '.docx':
            doc = Document(file_path)
            content = "\n".join([para.text for para in doc.paragraphs if para.text.strip()])
        elif ext == '.pdf':
            with open(file_path, 'rb') as f:
                reader = PyPDF2.PdfReader(f)
                content = "\n".join([page.extract_text() for page in reader.pages if page.extract_text()])
    except Exception as e:
        log_func(f"❌ 讀取文件 {os.path.basename(file_path)} 內容時失敗: {e}")
    return content


def init_browser(log_func=print, chrome_driver_path=None):
    """自動處理瀏覽器啟動邏輯與降級容錯"""
    log_func("🔍 嘗試啟動首選瀏覽器：Google Chrome...")
    try:
        chrome_options = Options()
        chrome_options.add_experimental_option("detach", True)
        chrome_options.add_argument('--ignore-certificate-errors')  # 👈 添加这一行即可无视警告 消失了
        chrome_options.add_argument('--disable-background-timer-throttling')
        chrome_options.add_argument('--disable-backgrounding-occluded-windows')
        chrome_options.add_argument('--disable-renderer-backgrounding')
        # chrome_options.add_argument('--headless=new') #当完全没有问题之后可以使用这个无头模式，不用盯着界面看，节省70%内存，可以调高worknum
        # chrome_options.add_argument('--disable-gpu')
        path = chrome_driver_path if chrome_driver_path else ChromeDriverManager().install()
        driver = webdriver.Chrome(service=Service(path), options=chrome_options)
        log_func("✅ 成功啟動 Google Chrome！")
        return driver
    except Exception as e_chrome:
        log_func(f"⚠️ Chrome 啟動失敗: {e_chrome}")
        log_func("🔄 正在嘗試啟動備用瀏覽器：Microsoft Edge...")
        try:
            edge_options = EdgeOptions()
            edge_options.add_experimental_option("detach", True)
            edge_options.add_argument('--ignore-certificate-errors')
            driver = webdriver.Edge(service=EdgeService(EdgeChromiumDriverManager().install()), options=edge_options)
            log_func("✅ 成功啟動 Microsoft Edge！")
            return driver
        except Exception as e_edge:
            log_func(f"❌ [致命錯誤] Chrome 和 Edge 均啟動失敗，請檢查瀏覽器安裝環境！\n錯誤資訊: {e_edge}")
            raise e_edge


def perform_login(driver, username, password, log_func=print):
    """封裝自動登錄流程"""
    log_func("🔑 正在執行自動登錄...")
    try:
        login_nav = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.ID, "nav-login")))
        driver.execute_script("arguments[0].click();", login_nav)

        username_input = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "sso-username")))
        password_input = driver.find_element(By.ID, "sso-password")

        driver.execute_script("arguments[0].value = arguments[1];", username_input, username)
        driver.execute_script("arguments[0].dispatchEvent(new Event('input', { bubbles: true }));", username_input)
        driver.execute_script("arguments[0].value = arguments[1];", password_input, password)
        driver.execute_script("arguments[0].dispatchEvent(new Event('input', { bubbles: true }));", password_input)

        submit_btn = driver.find_element(By.CLASS_NAME, "sso-submit-btn")
        driver.execute_script("arguments[0].click();", submit_btn)

        log_func("⏳ 正在等待接口校验账号密码...")
        time.sleep(1)  # 给前端 1 秒钟时间接收后端报错并渲染出粉红色的提示框

        # 使用 XPath 精确扫描页面上是否出现了截图中的两句报错文本
        error_elements = driver.find_elements(
            By.XPATH,
            "//*[contains(text(), 'Invalid password') or contains(text(), 'User does not exist')]"
        )

        # 遍历找到的元素，只要它是肉眼可见的，就说明登录被拒绝了
        for el in error_elements:
            if el.is_displayed():
                error_text = el.text.strip()
                raise ValueError(f"登录被系统拒绝，原因: {error_text}")

        close_popups(driver, log_func)
        log_func("✅ 登錄動作已提交，等待頁面刷新...")
        time.sleep(1)
        close_popups(driver, log_func)
    except Exception as e:
        log_func(f"❌ [登錄失敗] 請檢查頁面是否卡頓或元素發生變化: {e}")
        raise e


def get_react_time_extraction_js():
    """返回用於提取 DOM 準備與完成時間的龐大 JS 腳本字串"""
    return """
        try {
    var prep = 'N/A', comp = 'N/A';
    var debugLog = [];
    // 锁定我们要找的滚动组件标签
    var flows = document.querySelectorAll('number-flow-react, number-flow');
    debugLog.push('找到标签数: ' + flows.length);
    var vals = [];

    for(var i=0; i<flows.length; i++) {
        var el = flows[i];
        var val = null;

        // 第一层：尝试常规 DOM 属性
        if (el.value !== undefined) val = el.value;
        else if (el.getAttribute('value')) val = el.getAttribute('value');
        else if (el.getAttribute('aria-valuenow')) val = el.getAttribute('aria-valuenow');

        // 第二层：入侵 React 16-18+ 的 Fiber 内存树，向上回溯寻找 memoizedProps
        if (val == null) {
            var fiberKey = Object.keys(el).find(k => k.startsWith('__reactFiber$'));
            if (fiberKey) {
                var curr = el[fiberKey];
                // 向上回溯 5 层，寻找包含 value 变量的节点
                for (var j=0; j<5; j++) {
                    if (curr && curr.memoizedProps && curr.memoizedProps.value !== undefined) {
                        val = curr.memoizedProps.value;
                        debugLog.push('通过Fiber提取成功');
                        break;
                    }
                    if (curr) curr = curr.return;
                }
            }
        }

        // 第三层：入侵 React 17+ 的 Props 内存
        if (val == null) {
            var propsKey = Object.keys(el).find(k => k.startsWith('__reactProps$'));
            if (propsKey && el[propsKey] && el[propsKey].value !== undefined) {
                val = el[propsKey].value;
                debugLog.push('通过Props提取成功');
            }
        }

        if (val !== null) vals.push(val);
    }

    if (vals.length >= 1) prep = vals[0] + 's';
    if (vals.length >= 2) comp = vals[1] + 's';

    return { prep: prep, comp: comp, debug: debugLog.join(' | ') };
} catch(err) {
    return { prep: 'N/A', comp: 'N/A', debug: 'JS报错: ' + err.message };
}
    """