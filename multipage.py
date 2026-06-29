import os
import re
import time
from datetime import datetime
import threading
import queue
import openpyxl
import keyboard
from concurrent.futures import ThreadPoolExecutor
from webdriver_manager.chrome import ChromeDriverManager
from utils.logger import setup_global_logger
from utils.config import USERNAME, PASSWORD, HOME_URL, INPUT_EXCEL_PATH, MAX_WORKERS
from utils.common import (
    show_authorship,
    get_test_data_from_excel,
    init_browser,
    perform_login
)
from tools.summary_generator import generate_summary_csv

# ================= 引入我们打造的 Core 模块 =================
from utils.signals import STOP_EVENT
from utils.models import TaskInput
from utils.pipeline import process_single_task

PRINT_LOCK = threading.Lock()
CONSECUTIVE_CLOSES = 0
CLOSES_LOCK = threading.Lock()

def safe_print(*args, **kwargs):
    with PRINT_LOCK:
        print(*args, **kwargs)

def listen_for_hotkey():
    try:
        keyboard.add_hotkey('ctrl+alt+h', show_authorship)
        keyboard.wait('ctrl+q')
        STOP_EVENT.set()
        print("\n\n🛑 [紧急刹车] 侦测到 Ctrl+Q 键按下！脚本将在当前步骤跳出...\n")
    except ImportError:
        print("⚠️ 缺少 keyboard 模块，无法使用快捷键终止功能。")
    except Exception as e:
        print(f"⚠️ 热键监听启动失败 (可能缺少管理员权限)，将禁用 Ctrl+Q 功能: {e}")

def excel_writer_worker(excel_path, q):
    """专职 Excel 写入后台线程"""
    import openpyxl
    try:
        wb = openpyxl.load_workbook(excel_path)
        ws = wb.active
    except Exception as e:
        safe_print(f"❌ [写线程致命错误] 无法加载 Excel: {e}")
        return

    while True:
        item = q.get()
        if item == "STOP":
            try: wb.save(excel_path)
            except: pass
            break

        task, result = item
        try:
            row_data = [
                task.index + 1, task.question_text, task.filename if task.filename else "",
                result.crash_reason, result.tester_expectation, task.target_language if task.target_language else "N/A",
                result.input_language, result.output_language, result.language_status,
                result.answer_text, result.shared_link, result.evaluation_text,
                result.actual_agent_used, result.reference_link, result.document_contain,
                result.prep_time, result.comp_time, result.timeout_status
            ]
            target_row = task.index + 2
            for col_index, value in enumerate(row_data, start=1):
                if isinstance(value, str):
                    # 过滤掉 Excel 不支持的底层控制字符
                    value = re.sub(r'[\x00-\x08\x0b\x0c\x0e-\x1f]', '', value)
                ws.cell(row=target_row, column=col_index, value=value)

            save_success = False
            for retry in range(3):
                try:
                    wb.save(excel_path)
                    save_success = True
                    break
                except PermissionError:
                    time.sleep(3)
            if not save_success:
                safe_print(f"❌ [致命错误] 多次尝试保存失败，第 {task.index + 1} 行结果丢失！")
        except Exception as e:
            safe_print(f"❌ [写线程异常] 写入行 {task.index + 1} 失败: {e}")
        finally:
            q.task_done()

def worker_thread(task_index, task_input: TaskInput, test_dir, result_queue, browser_pool, cached_driver_path):
    """
    独立线程工作者：只负责排队借用浏览器、执行流水线、然后归还浏览器。
    """
    if STOP_EVENT.is_set():
        return

    log_func = lambda msg: safe_print(f"[{task_index + 1}] {msg}")
    driver = None

    # 1. 从资源池获取闲置浏览器
    log_func("⏳ 正在等待分配闲置浏览器...")
    while True:
        if STOP_EVENT.is_set(): return
        try:
            driver = browser_pool.get(timeout=3)
            break
        except queue.Empty:
            continue

    # 2. 清理浏览器残余状态 (暴力兜底弹窗)
    try:
        try:
            from selenium.webdriver.support.ui import WebDriverWait
            from selenium.webdriver.support import expected_conditions as EC
            WebDriverWait(driver, 0.5).until(EC.alert_is_present())
            driver.switch_to.alert.accept()
        except Exception:
            pass

        try:
            driver.get(HOME_URL)
            if STOP_EVENT.wait(timeout=1.5):return
        except Exception as e:
            log_func(f"⚠️ 浏览器重置失败，可能已损坏: {e}")

    # ==============================================================
    # 3. 核心魔法：一行代码完成所有业务逻辑！
        status = process_single_task(
            task=task_input,
            driver=driver,
            test_dir=test_dir,
            result_queue=result_queue,
            stop_event=STOP_EVENT,
            log_func=log_func
        )
    # ==============================================================

    # 4. 全局阻断机制与现场保留
        global CONSECUTIVE_CLOSES

        if status == "BROWSER_CLOSED":
            with CLOSES_LOCK:
                CONSECUTIVE_CLOSES += 1
                if CONSECUTIVE_CLOSES >= MAX_WORKERS and not STOP_EVENT.is_set():
                    safe_print("\n🛑 [全局拦截] 侦测到一口气关闭了所有窗口，判定为终止指令，脚本结束！")
                    STOP_EVENT.set()

        elif status == "SUCCESS":
            with CLOSES_LOCK:
                CONSECUTIVE_CLOSES = 0  # 只要有成功跑完的，就重置计数器

        elif status == "ERROR" and not STOP_EVENT.is_set():
            with CLOSES_LOCK:
                CONSECUTIVE_CLOSES = 0  # 普通报错也重置，防止误判
            log_func("⏸️ 任务报错，保留浏览器现场 20 秒供排查...")
            if STOP_EVENT.wait(timeout=20): return
    except Exception as worker_err:
        log_func(f"❌ [严重警报] 线程内部发生未捕获的异常，被成功拦截: {worker_err}")
    finally:
    # 5. 归还或重建损毁的浏览器
        if driver:
            try:
                _ = driver.title  # 试探性获取 title，报错说明被强关或崩溃
                browser_pool.put(driver)
            except Exception:
                if not STOP_EVENT.is_set():
                    log_func("🔄 检测到浏览器已损毁，正在为您创建替补浏览器并放入池中...")
                    try:
                        if driver: driver.quit()
                    except Exception as quit_err:
                        log_func(f"🧹 浏览器销毁时出现预期内的警告 (已忽略): {quit_err}")
                    try:
                        new_driver = init_browser(log_func=lambda msg: None, chrome_driver_path=cached_driver_path)
                        new_driver.get(HOME_URL)
                        time.sleep(1)
                        perform_login(new_driver, USERNAME, PASSWORD, log_func=lambda msg: None)
                        browser_pool.put(new_driver)
                    except Exception as rebuild_err:
                        log_func(f"❌ 创建替补浏览器彻底失败: {rebuild_err}")
                        log_func("🚨 触发防死锁机制，全局终止自动化任务！")
                        STOP_EVENT.set()  # 👈 新增这一行，发出全局刹车信号


def run_automation():
    setup_global_logger()
    print("📥 正在初始化並緩存瀏覽器驅動...")
    cached_driver_path = ChromeDriverManager().install()
    threading.Thread(target=listen_for_hotkey, daemon=True).start()

    project_dir = os.path.dirname(os.path.abspath(__file__))
    test_dir = os.path.join(project_dir, "test")

    if not os.path.exists(test_dir):
        print(f"⚠️ [警告] 找不到测试文件夹: {test_dir}，已自动创建。若本次测试包含文件上传，请确保文件放入该目录。")
        os.makedirs(test_dir, exist_ok=True)

    questions, selected_agents, target_filenames, selected_languages = get_test_data_from_excel(INPUT_EXCEL_PATH, log_func=safe_print)
    if not questions:
        print("未提取到任何问题，程序终止。")
        return
    # ================= 文件数量校验 =================
    valid_extensions = ('.pdf', '.docx', '.csv', '.txt')
    test_files = [f for f in os.listdir(test_dir) if f.lower().endswith(valid_extensions)]
    expected_file_count = len([f for f in target_filenames if str(f).strip()])
    if len(test_files) != expected_file_count:
        print(f"⚠️ 文件夹中的文件数量 ({len(test_files)}) 与 Excel 中的问题数量 ({len(questions)}) 不匹配！")
    # ================= 核心修改：打包 TaskInput =================
    task_inputs = []
    for i in range(len(questions)):
        task_inputs.append(TaskInput(
            index=i,
            question_text=questions[i],
            target_agent=selected_agents[i],
            filename=target_filenames[i],
            target_language=selected_languages[i]
        ))

    # ================= 初始化 Excel 文件 (同原版) =================
    base_testcase_name = os.path.splitext(os.path.basename(INPUT_EXCEL_PATH))[0]
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    dynamic_filename = f"evaluation_results_{base_testcase_name}_{timestamp}.xlsx"
    excel_dir = os.path.join(project_dir, "Evaluation_Results")
    os.makedirs(excel_dir, exist_ok=True)
    excel_path = os.path.join(excel_dir, dynamic_filename)

    if not os.path.exists(excel_path):
        wb = openpyxl.Workbook()
        wb.properties.creator = "Henry HONG "
        wb.properties.description = "Authored by Henry HONG. "
        ws = wb.active
        ws.title = "Evaluation Results"
        ws.append(["label", "Request", "filename", "Crash", "Tester Expectation", "Selected Language", "Input Language",
                   "Output Language", "Language Overall Status", "answer", "shared link", "DeepSeek评价内容",
                   "Selected agent", "Reference Link", "Document Contain[1][2][3]", "Preparation Time",
                   "Completion Time", "Timeout_States"])
        wb.save(excel_path)
        print(f"📊 已创建评价结果 Excel 文件: {excel_path}")

    # ================= 预热浏览器池 (同原版) =================
    result_queue = queue.Queue()
    writer_thread = threading.Thread(target=excel_writer_worker, args=(excel_path, result_queue), daemon=True)
    writer_thread.start()
    print(f"\n♨️ 正在预热浏览器池，将提前启动并登录 {MAX_WORKERS} 个浏览器，请稍候...")
    browser_pool = queue.Queue()

    for init_i in range(MAX_WORKERS):
        if STOP_EVENT.is_set(): break
        d = None
        try:
            d = init_browser(log_func=lambda msg: safe_print(f"[预热 {init_i + 1}/{MAX_WORKERS}] {msg}"),
                             chrome_driver_path=cached_driver_path)
            d.get(HOME_URL)
            time.sleep(1)
            perform_login(d, USERNAME, PASSWORD, log_func=lambda msg: safe_print(f"[预热 {init_i + 1}/{MAX_WORKERS}] {msg}"))
            browser_pool.put(d)
        except Exception as e:
            print(f"\n❌ [致命错误] 预热第 {init_i + 1} 个浏览器时启动或登录失败: {e}")
            try:
                if d: d.quit()
            except:
                pass
            while not browser_pool.empty():
                try:
                    browser_pool.get_nowait().quit()
                except:
                    pass
            return
    if browser_pool.empty():
        print("\n❌ [致命错误] 所有浏览器预热均失败，池中无可用浏览器！")
        return


    # ================= 并发派发任务 =================
    print(f"\n🚀 预热完毕！启动并发模式...")
    try:
        with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
            for task in task_inputs:
                if STOP_EVENT.is_set(): break
                executor.submit(
                    worker_thread,
              task.index, task, test_dir, result_queue, browser_pool, cached_driver_path
                )
    except KeyboardInterrupt:
        print("\n🛑 侦测到系统强行中断信号 (Ctrl+C)！正在紧急通知所有并行浏览器停止工作...")
        STOP_EVENT.set()
    except Exception as e:
        print(f"\n❌ 主脚本发生致命崩溃: {e}")
    finally:
    # ================= 生成报告与清理  =================
        print("💾 正在等待残余数据写入 Excel...")
        result_queue.put("STOP")
        writer_thread.join(timeout=10)
        print("📊 正在生成最终的 Summary 报告...")
        output_csv = os.path.join(project_dir, "Summaries", f"Summary_{base_testcase_name}_{timestamp}.csv")
        os.makedirs(os.path.dirname(output_csv), exist_ok=True)
        try:
            generate_summary_csv(excel_path, output_csv)
            print(f"✅ 汇总报告已生成: {output_csv}")
        except Exception as e:
            print(f"⚠️ 生成汇总报告时发生异常: {e}")

        if STOP_EVENT.is_set():
            print("\n🛑 任务已被手动中断！")
        else:
            print("\n✅ 文件夹内所有测试用例已全部运行完毕！")

        print("🧹 正在清理浏览器池释放内存...")
        while not browser_pool.empty():
            try:
                d = browser_pool.get_nowait()
                d.quit()
            except: pass

        print("👉 自动化流程结束，控制台即将退出。")

if __name__ == "__main__":
    run_automation()
